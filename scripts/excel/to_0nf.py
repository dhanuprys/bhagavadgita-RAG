import json

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Load JSON from file
def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# Load data from the specified paths
chapters = load_json("data/3-fine-verse_number/chapters.json")
verses = load_json("data/3-fine-verse_number/verses.json")
translations = load_json("data/3-fine-verse_number/translations.json")

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Data Gita 0NF"

# Headers (Bahasa + technical name)
headers = [
    "Nomor Bab (chapter_number)",
    "Nama Bab (name)",
    "Nama Bab (Hindi) (name_hindi)",
    "Nama Bab (Sanskrit) (name_sanskrit)",
    "Ringkasan Bab (summary)",
    "Jumlah Ayat (verses_count)",
    "Nomor Ayat (verse_number)",
    "Teks Hindi (text_hindi)",
    "Teks Sanskrit (text_sanskrit)",
    "Arti Per Kata (Sanskrit) (text_sanskrit_meanings)",
    "URL Audio (audio_url)",
    "Terjemahan Ayat (content)",
]
ws.append(headers)

# Track max width of each column
max_lengths = [len(h) for h in headers]

# Build a quick lookup for chapters by id
chap_by_id = {c["id"]: c for c in chapters}

# Group translations by verse_id
trans_by_verse = {}
for t in translations:
    trans_by_verse.setdefault(t["verse_id"], []).append(t["content"])

# Populate rows
for verse in verses:
    chap = chap_by_id.get(verse["chapter_id"], {})
    verse_id = verse.get("id")
    matched_trans = trans_by_verse.get(verse_id, [])

    # if no translations, still write one empty row
    if not matched_trans:
        rows = [None]
    else:
        rows = matched_trans

    for content in rows:
        row = [
            chap.get("chapter_number", ""),
            chap.get("name", ""),
            chap.get("name_hindi", ""),
            chap.get("name_sanskrit", ""),
            chap.get("summary", ""),
            chap.get("verses_count", ""),
            verse.get("verse_number", ""),
            verse.get("text_hindi", ""),
            verse.get("text_sanskrit", ""),
            verse.get("text_sanskrit_meanings", ""),
            verse.get("audio_url", ""),
            content or "",
        ]
        ws.append(row)
        # update max lengths
        for i, cell in enumerate(row):
            l = len(str(cell))
            if l > max_lengths[i]:
                max_lengths[i] = l

# Adjust column widths
for i, width in enumerate(max_lengths, start=1):
    col = get_column_letter(i)
    ws.column_dimensions[col].width = min(width + 2, 80)

# Save to the desired output path
wb.save("data/gita_0NF.xlsx")
print("✅ data/gita_0NF.xlsx written, and Terjemahan Ayat is now populated correctly.")
